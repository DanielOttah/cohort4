from openpyxl import Workbook, load_workbook


# HARD CODING THE HEADERS
def build_dict():
    wkBk = load_workbook("./excel_files/Invoices_App.xlsx")
    db = {}
    tables = {}
    sheets = []
    hdrs = []
    details = {}
    for tabs in wkBk.sheetnames:
        sheets.append(tabs)

    for row in wkBk[sheets[0]].iter_rows(min_row=2, max_row=wkBk[sheets[0]].max_row, values_only=True):
        customer_id = row[0]
        details = {
            'first_name': row[1], 'last_name': row[2], 'email': row[3], 'phone': row[4]}
        tables[customer_id] = details
        db[sheets[0]] = tables
    tables = {}
    for row in wkBk[sheets[1]].iter_rows(min_row=2, max_row=wkBk[sheets[1]].max_row, values_only=True):
        invoice_id = row[0]
        details = {
            'customer_id': row[1], 'invoice_number': row[2], 'date_of_invoice': row[3]}
        tables[invoice_id] = details
        db[sheets[1]] = tables
    tables = {}
    for row in wkBk[sheets[2]].iter_rows(min_row=2, max_row=wkBk[sheets[2]].max_row, values_only=True):
        invoice_list_id = row[0]
        details = {
            'invoice_id': row[1], 'product_id': row[2], 'invoice_qty': row[3]}
        tables[invoice_list_id] = details
        db[sheets[2]] = tables
    tables = {}
    for row in wkBk[sheets[3]].iter_rows(min_row=2, max_row=wkBk[sheets[3]].max_row, values_only=True):
        product_id = row[0]
        details = {
            'product_name': row[1], 'price': row[2]}
        tables[product_id] = details
        db[sheets[3]] = tables
    tables = {}
    return db

# ANOTHER METHOD


def breakIntoDict():
    wb = load_workbook("./excel_files/Invoices_App.xlsx")
    builtDict = {}
    for sheet in wb.sheetnames:
        builtDict[sheet] = {}
        rowCount = 1
        for row in wb[sheet].iter_rows(min_row=2, min_col=1):
            if row[0].value is not None:
                col = 0
                builtDict[sheet][rowCount] = {}
                for cell in row:
                    builtDict[sheet][rowCount][wb[sheet]
                                               [1][col].value] = cell.value
                    col += 1
            rowCount += 1
    # print(builtDict)
    return builtDict


print(breakIntoDict())
