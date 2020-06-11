from openpyxl import Workbook, load_workbook

wkBk = load_workbook("./excel_files/Invoices_App.xlsx")


def getInvoice(invoice_id):
    inv_sheet = wkBk['Invoices']
    for line in inv_sheet['A']:
        if line.value == invoice_id:
            # returns customerID and invoice number
            return (inv_sheet.cell(row=line.row, column=line.column+1).value,
                    inv_sheet.cell(row=line.row, column=line.column+2).value)


def getCustomer(customer_id):
    cust_sheet = wkBk['customers']
    for line in cust_sheet['A']:
        if line.value == customer_id:
            fname = cust_sheet.cell(row=line.row, column=line.column+1).value
            lname = cust_sheet.cell(row=line.row, column=line.column+2).value
            email = cust_sheet.cell(row=line.row, column=line.column+3).value
            phone = cust_sheet.cell(row=line.row, column=line.column+4).value
            customer = Customer(fname, lname, email, phone)
            return customer


def getInvoiceList(invoice_id):
    invList_sheet = wkBk['Invoice_list']
    for line in invList_sheet['B']:
        if line.value == invoice_id:
            # returns products bought and quantity
            return (invList_sheet.cell(row=line.row, column=line.column+1).value,
                    invList_sheet.cell(row=line.row, column=line.column+2).value)


def getProductBought(prod_id):
    prod_sheet = wkBk['products']
    for line in prod_sheet['A']:
        if line.value == prod_id:
            # returns products bought
            return (prod_sheet.cell(row=line.row, column=line.column+1).value,
                    prod_sheet.cell(row=line.row, column=line.column+2).value)


class Customer():
    def __init__(self, fname, lname, email, phone):
        self.fname = fname
        self.lname = lname
        self.email = email
        self.phone = phone

    def customerDetail(self):
        return {'name': f"{self.fname} {self.lname}", 'email': self.email, 'phone': self.phone}

    def customerFullName(self):
        return f"{self.fname} {self.lname}"


def report_Invoice(invoice_id):
    cust = getInvoice(invoice_id)  # nb cust is (customer_id,invoice number)
    customer = getCustomer(cust[0])
    goods = getInvoiceList(invoice_id)  # nb goods is (product_id,quantity)
    item = getProductBought(goods[0])  # nb item is (product,price)
    total_item_cost = item[1]*goods[1]
    return f"""
    Invoice number: {cust[1]}
    Customer name: {customer.customerFullName()}
    Customer email: {customer.email}
    Customer phone: {customer.phone}
    Item(s): {item[0]}
    Quantity: {goods[1]}
    Total cost: $CAD {total_item_cost}"""
