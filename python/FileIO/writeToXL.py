import openpyxl
from food import results

# API for the food list - http://tropicalfruitandveg.com/api/tfvjsonapi.php?search=all


wb = openpyxl.Workbook()  # Create Workbook

sheet = wb.active  # get active sheet from workbook

sheet_title = sheet.title
# print(sheet_title)
sheet.title = "food"
# print(results[0])

# c1 = sheet.cell(row=1, column=1)
# c1.value = "ANKIT"
# wb['food'].cell(row=2, column=2).value = "Ottah"

# Create the headers in Excel
rowCount = 1
for key in results[0]:
    sheet.cell(row=1, column=rowCount).value = key
    rowCount += 1

# Enter the values as rows
rowCnt = 2
for eachDict in results:
    colCnt = 1
    for key, value in eachDict.items():
        sheet.cell(row=rowCnt, column=colCnt).value = value
        colCnt += 1
    rowCnt += 1


wb.save('C:\\code\\cohort4\\python\\FileIO\\food.xlsx')
